"""
Parsers for the two source workbook formats used by the sites:

1. Weekly Coal Sales Report  ->  sheet "Stock Inventory"
   Fixed layout (confirmed against real files):
     Row 7  = Crushed Coal   | Opening=E, Stock In=H, Stock Out=J, Balance=L
     Row 8  = Uncrushed Coal | Opening=E, Stock In=H, Stock Out=J, Balance=L
     Row 9  = Coal Dust      | Opening=E, Stock In=H, Stock Out=J, Balance=L

2. Diesel Dispense report (daily log + a pivot summary sheet)
   The pivot sheet has a "Row Labels" header cell, one row per operation
   type, and a numeric total in the right-most populated column of that
   row (works for both the TRCM layout - a single total column - and the
   IFCM layout - one column per day plus a "Grand Total" column).
"""

from typing import Optional

import openpyxl
import pandas as pd


def parse_coal_stock_report(file) -> dict:
    """Return {"Crushed Coal": {...}, "Uncrushed Coal": {...}, "Coal Dust": {...}}
    Each value dict has opening/stock_in/stock_out/balance.
    Raises ValueError with a clear message if the expected sheet/layout isn't found.
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    if "Stock Inventory" not in wb.sheetnames:
        raise ValueError(
            "Couldn't find a 'Stock Inventory' sheet in this workbook. "
            f"Sheets found: {wb.sheetnames}"
        )
    ws = wb["Stock Inventory"]

    expected_rows = {7: "Crushed Coal", 8: "Uncrushed Coal", 9: "Coal Dust"}
    result = {}
    for row_num, expected_name in expected_rows.items():
        item_name = ws.cell(row=row_num, column=3).value  # column C
        if not item_name or expected_name.split()[0].lower() not in str(item_name).lower():
            raise ValueError(
                f"Expected '{expected_name}' in row {row_num} (cell C{row_num}) "
                f"but found '{item_name}'. The report layout may have changed - "
                "please check the file or adjust the row mapping in parsers.py."
            )
        result[expected_name] = {
            "opening": ws.cell(row=row_num, column=5).value,   # E
            "stock_in": ws.cell(row=row_num, column=8).value,  # H
            "stock_out": ws.cell(row=row_num, column=10).value,  # J
            "balance": ws.cell(row=row_num, column=12).value,  # L
        }
    return result


def parse_diesel_pivot(file) -> pd.DataFrame:
    """Scan every sheet for a pivot-style summary (a cell containing
    'Row Labels') and return a DataFrame [operation_type, total_litres]
    for the first one found. Skips the 'Grand Total' row itself.
    """
    wb = openpyxl.load_workbook(file, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_cell = None
        for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row)):
            for cell in row:
                if isinstance(cell.value, str) and "row labels" in cell.value.lower():
                    header_cell = cell
                    break
            if header_cell:
                break
        if header_cell is None:
            continue

        header_row = header_cell.row
        label_col = header_cell.column

        total_col = label_col + 1
        for cell in ws[header_row]:
            if isinstance(cell.value, str) and "grand total" in cell.value.lower():
                total_col = cell.column
                break

        data = []
        r = header_row + 1
        blank_streak = 0
        while r <= ws.max_row:
            label = ws.cell(row=r, column=label_col).value
            if label is None:
                blank_streak += 1
                if blank_streak > 2:
                    break
                r += 1
                continue
            blank_streak = 0
            label_str = str(label).strip()
            if label_str.lower() == "grand total":
                break
            if label_str.lower() in ("(blank)", "stock in", ""):
                r += 1
                continue
            total = ws.cell(row=r, column=total_col).value
            if isinstance(total, (int, float)):
                data.append({"operation_type": label_str, "total_litres": total})
            r += 1

        if data:
            return pd.DataFrame(data)

    raise ValueError(
        "Couldn't find a pivot summary (a 'Row Labels' table) in this workbook. "
        f"Sheets found: {wb.sheetnames}"
    )


def summarize_diesel_by_category(pivot_df: pd.DataFrame, site: str, classify_fn) -> dict:
    """Split a parsed diesel pivot into Haulage vs Non-Haulage totals using the
    provided classify_fn(site, operation_type) -> 'Haulage' | 'Non-Haulage'.
    Returns {"haulage": total, "non_haulage": total, "detail": {op: category}}.
    """
    haulage = 0.0
    non_haulage = 0.0
    detail = {}
    for _, row in pivot_df.iterrows():
        cat = classify_fn(site, row["operation_type"])
        detail[row["operation_type"]] = cat
        if cat == "Haulage":
            haulage += float(row["total_litres"])
        else:
            non_haulage += float(row["total_litres"])
    return {"haulage": haulage, "non_haulage": non_haulage, "detail": detail}
